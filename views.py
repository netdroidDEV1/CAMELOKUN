from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import EncuestaExperiencia


from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .models import EncuestaExperiencia

@require_http_methods(["GET", "POST"])
def encuesta_create(request):
    """
    Vista principal para crear una encuesta pública.
    Los usuarios no autenticados pueden acceder.
    Los usuarios autenticados también pueden acceder (verán el menú si el template lo permite).
    """
    if request.method == 'GET':
        return render(request, 'encuestas/index.html')

    data = request.POST
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    errors = {}

    # ✅ Opciones válidas para cada campo con selección predefinida
    required_choices = {
        'eps': ['Cajacopi', 'Dusakawi', 'Nueva EPS'],
        'asignacion_cita': ['1 a 3 días', '4 a 6 días', '7 a 9 días', 'Más de 10 días'],
        'instalaciones_seguridad': ['EXCELENTE', 'BUENO', 'REGULAR', 'MALO'],
        'atencion_personal_admin': ['EXCELENTE', 'BUENO', 'REGULAR', 'MALO'],
        'atencion_profesional_salud': ['EXCELENTE', 'BUENO', 'REGULAR', 'MALO'],
        'experiencia_global': ['EXCELENTE', 'BUENO', 'REGULAR', 'MALO'],
        'recomendaria_ipsi': [
            'DEFINITIVAMENTE SÍ',
            'PROBABLEMENTE SÍ',
            'DEFINITIVAMENTE NO',
            'PROBABLEMENTE NO'
        ],
    }

    # 🔍 Validación de campos requeridos
    for field, opciones in required_choices.items():
        valor = data.get(field)
        if not valor:
            errors[field] = 'Este campo es obligatorio.'
        elif valor not in opciones:
            errors[field] = f'Valor inválido: {valor}.'

    # Validar campo libre “servicio”
    if not data.get('servicio'):
        errors['servicio'] = 'Por favor indique el servicio al que asistió.'

    if errors:
        if is_ajax:
            return JsonResponse({'ok': False, 'errors': errors}, status=400)
        return render(request, 'encuestas/index.html', {'errors': errors, 'old': data})

    # 💾 Guardar en la base de datos
    encuesta = EncuestaExperiencia.objects.create(
        numero_identificacion=data.get('numero_identificacion') or None,
        eps=data.get('eps'),
        servicio=data.get('servicio'),
        asignacion_cita=data.get('asignacion_cita'),
        instalaciones_seguridad=data.get('instalaciones_seguridad'),
        atencion_personal_admin=data.get('atencion_personal_admin'),
        atencion_profesional_salud=data.get('atencion_profesional_salud'),
        experiencia_global=data.get('experiencia_global'),
        recomendaria_ipsi=data.get('recomendaria_ipsi'),
        mejoras_sugeridas=data.get('mejoras_sugeridas') or None,
    )

    if is_ajax:
        return JsonResponse({'ok': True, 'id': encuesta.id, 'redirect': reverse('encuesta_gracias')})

    return redirect('encuesta_gracias')


def encuesta_gracias(request):
    """Página de agradecimiento tras enviar la encuesta."""
    return render(request, 'encuestas/gracias.html')


from django.shortcuts import render, redirect
from django.contrib import messages
from encuestas.models import UsuarioEncuesta

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        try:
            usuario = UsuarioEncuesta.objects.get(username=username)
            if usuario.password == password:
                request.session['usuario_id'] = usuario.id
                request.session['usuario_username'] = usuario.username
                return redirect('index')
            else:
                messages.error(request, 'Contraseña incorrecta.')
        except UsuarioEncuesta.DoesNotExist:
            messages.error(request, 'Usuario no encontrado.')

    return render(request, 'login.html')


def index_view(request):
    """
    Vista pública del formulario principal (index).
    - Usuarios no autenticados: ven el formulario de encuesta normalmente.
    - Usuarios autenticados: también pueden verlo, mostrando su nombre si se desea.
    """
    usuario = request.session.get('usuario_username')  # opcional
    return render(request, 'index.html', {'usuario': usuario})


from django.shortcuts import redirect

def logout_view(request):
    """Cierra la sesión actual y redirige al login."""
    request.session.flush()  # elimina todos los datos de sesión
    return redirect('login')

import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import EncuestaExperiencia


def exportar_excel(request):
    """
    Exporta todas las encuestas en formato Excel (.xlsx)
    con formato profesional, encabezados filtrables, y nuevos campos EPS y Servicio.
    """

    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Encuestas IPSI Wayuu Talatshi"

    # === Encabezados ===
    encabezados = [
        'ID',
        'Fecha de creación',
        'Número de identificación',
        'EPS',
        'Servicio',
        'Asignación de cita',
        'Instalaciones y seguridad',
        'Atención personal administrativo',
        'Atención profesional de salud',
        'Experiencia global',
        'Recomendación IPSI',
        'Mejoras sugeridas',
    ]
    ws.append(encabezados)

    # === Estilo de encabezados ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="006C84", end_color="006C84", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === Agregar datos ===
    for encuesta in EncuestaExperiencia.objects.all().order_by('-fecha_creacion'):
        ws.append([
            encuesta.id,
            encuesta.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            encuesta.numero_identificacion or '',
            encuesta.eps,
            encuesta.servicio,
            encuesta.asignacion_cita,
            encuesta.instalaciones_seguridad,
            encuesta.atencion_personal_admin,
            encuesta.atencion_profesional_salud,
            encuesta.experiencia_global,
            encuesta.recomendaria_ipsi,
            encuesta.mejoras_sugeridas or '',
        ])

    # === Autoajuste de columnas ===
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # === Activar filtros automáticos ===
    ws.auto_filter.ref = f"A1:L{ws.max_row}"  # Aplica filtros a todas las columnas

    # === Congelar encabezado ===
    ws.freeze_panes = "A2"  # Mantiene la fila de encabezado fija al hacer scroll

    # === Preparar respuesta HTTP ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Encuestas_IPSI_{datetime.date.today()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response



import io
import datetime
import matplotlib.pyplot as plt
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .models import EncuestaExperiencia


def analisis_pdf(request):
    """
    Genera un informe PDF profesional con análisis estadístico,
    gráficos de satisfacción y distribución por EPS y Servicio.
    """

    # === Inicialización del documento ===
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=60,
        bottomMargin=40
    )
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='SubTitle', fontSize=11, textColor='#444'))
    styles.add(ParagraphStyle(name='Justify', alignment=4, leading=14))

    # === Encabezado ===
    title = Paragraph(
        "<b>📊 Informe de Análisis de Satisfacción del Usuario</b><br/>"
        "IPSI Wayuu Talatshi", styles["Title"]
    )
    subtitle = Paragraph(
        f"Generado automáticamente el {datetime.date.today().strftime('%d/%m/%Y')}<br/>"
        "<i>Área de Calidad y Experiencia del Usuario</i>",
        styles["SubTitle"]
    )

    elements += [title, subtitle, Spacer(1, 20)]

    encuestas = EncuestaExperiencia.objects.all()
    total = encuestas.count()

    if total == 0:
        elements.append(Paragraph("⚠️ No hay datos disponibles para generar el análisis.", styles["Normal"]))
        doc.build(elements)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="Analisis_Encuestas.pdf")

    # === 1️⃣ Resumen general de experiencia ===
    resumen_experiencia = {
        'EXCELENTE': encuestas.filter(experiencia_global="EXCELENTE").count(),
        'BUENO': encuestas.filter(experiencia_global="BUENO").count(),
        'REGULAR': encuestas.filter(experiencia_global="REGULAR").count(),
        'MALO': encuestas.filter(experiencia_global="MALO").count(),
    }
    total_respuestas = sum(resumen_experiencia.values())

    # === 2️⃣ Análisis por EPS ===
    resumen_eps = {}
    for eps in ["Cajacopi", "Dusakawi", "Nueva EPS"]:
        resumen_eps[eps] = encuestas.filter(eps=eps).count()

    # === 3️⃣ Análisis por servicio más frecuente ===
    servicios_top = {}
    for encuesta in encuestas:
        if encuesta.servicio:
            key = encuesta.servicio.strip().title()
            servicios_top[key] = servicios_top.get(key, 0) + 1
    servicios_top = dict(sorted(servicios_top.items(), key=lambda x: x[1], reverse=True)[:5])

    # === 4️⃣ Gráfico de satisfacción global ===
    plt.figure(figsize=(5, 3))
    plt.bar(
        resumen_experiencia.keys(),
        resumen_experiencia.values(),
        color=['#00a65a', '#007d46', '#ffb347', '#e74c3c']
    )
    plt.title('Distribución de la Experiencia Global')
    plt.xlabel('Nivel de satisfacción')
    plt.ylabel('Cantidad de respuestas')
    plt.tight_layout()

    graph_buffer = io.BytesIO()
    plt.savefig(graph_buffer, format='png')
    plt.close()
    graph_buffer.seek(0)

    elements.append(Image(graph_buffer, width=400, height=250))
    elements.append(Spacer(1, 20))

    # === 5️⃣ Tabla resumen general ===
    data = [["Nivel", "Cantidad", "Porcentaje"]]
    for nivel, cantidad in resumen_experiencia.items():
        porcentaje = (cantidad / total_respuestas * 100) if total_respuestas > 0 else 0
        data.append([nivel, cantidad, f"{porcentaje:.1f}%"])

    table = Table(data, hAlign='LEFT', colWidths=[150, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#006c84")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f8")),
        ('GRID', (0, 0), (-1, -1), 1, colors.gray),
    ]))
    elements += [Paragraph("<b>Resumen de Satisfacción Global</b>", styles["Heading2"]), table, Spacer(1, 20)]

    # === 6️⃣ Tabla de EPS ===
    data_eps = [["EPS", "Cantidad", "Porcentaje"]]
    for eps, cantidad in resumen_eps.items():
        porcentaje = (cantidad / total * 100)
        data_eps.append([eps, cantidad, f"{porcentaje:.1f}%"])

    table_eps = Table(data_eps, hAlign='LEFT', colWidths=[150, 100, 100])
    table_eps.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#006c84")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f8")),
        ('GRID', (0, 0), (-1, -1), 1, colors.gray),
    ]))
    elements += [Paragraph("<b>Distribución por EPS</b>", styles["Heading2"]), table_eps, Spacer(1, 20)]

    # === 7️⃣ Servicios más registrados ===
    if servicios_top:
        data_serv = [["Servicio", "Cantidad"]]
        for servicio, cantidad in servicios_top.items():
            data_serv.append([servicio, cantidad])

        table_serv = Table(data_serv, hAlign='LEFT', colWidths=[250, 100])
        table_serv.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#006c84")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f8")),
            ('GRID', (0, 0), (-1, -1), 1, colors.gray),
        ]))
        elements += [Paragraph("<b>Servicios más frecuentes</b>", styles["Heading2"]), table_serv, Spacer(1, 20)]

    # === 8️⃣ Conclusiones automáticas ===
    conclusion = ""
    if resumen_experiencia['EXCELENTE'] + resumen_experiencia['BUENO'] >= total_respuestas * 0.8:
        conclusion = "✅ Los resultados reflejan una alta satisfacción general. Se evidencia un ambiente de servicio positivo."
    elif resumen_experiencia['REGULAR'] >= total_respuestas * 0.3:
        conclusion = "⚠️ Se identifican oportunidades de mejora en atención y tiempos de respuesta. Reforzar capacitaciones y comunicación interna."
    else:
        conclusion = "❌ Se recomienda una revisión urgente del proceso de atención, priorizando acciones correctivas inmediatas."

    elements += [
        Paragraph("<b>Conclusión del Análisis</b>", styles["Heading2"]),
        Paragraph(conclusion, styles["Justify"]),
        Spacer(1, 20),
        Paragraph("<i>Área de Calidad - IPSI Wayuu Talatshi</i><br/>"
                  "Este informe ha sido generado automáticamente por el sistema de encuestas institucional.",
                  styles["Italic"]),
    ]

    # === Generar PDF ===
    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="Analisis_Encuestas.pdf")
